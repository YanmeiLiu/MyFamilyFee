import os

import openpyxl
import xlrd


class OptionExcel(object):
    def __init__(self, excelFile, excelPath=''):
        self.excelFile = excelFile
        self.excelPath = excelPath
        self.paralist = []
        self.caselist = []

    def getExcelData(self, excelFile, excelPath=''):
        xo = xlrd.open_workbook(excelFile)

        sheet = xo.sheet_by_name('case')
        col_num = sheet.ncols
        row_num = sheet.nrows
        # print('统计有%d条测试用例'%(row_num - 1))
        for n in range(1,row_num):

            tmplist = {}
            tmplist['id'] = n
            tmplist['business'] = sheet.cell(n,1).value
            tmplist['is_iqiyi'] = sheet.cell(n,2).value
            tmplist['is_video_page'] = sheet.cell(n,3).value
            tmplist['categoryid'] = sheet.cell(n,4).value
            tmplist['qypid'] = sheet.cell(n,5).value
            tmplist['expect'] = sheet.cell(n,6).value
            self.caselist.append(tmplist)
        return self.caselist

    def getParams(self, excelFile, excelPath=''):
        xo = xlrd.open_workbook(excelFile)

        sheet = xo.sheet_by_name('case')
        col_num = sheet.ncols
        row_num = sheet.nrows
        # print('统计有%d条有效用例' % (row_num - 1))
        for n in range(1, row_num):
            tmplist = {}
            tmplist['business'] = sheet.cell(n, 1).value
            tmplist['iqiyi'] = sheet.cell(n, 2).value
            tmplist['is_video_page'] = sheet.cell(n, 3).value
            tmplist['categoryid'] = sheet.cell(n, 4).value
            tmplist['qypid'] = sheet.cell(n, 5).value
            self.paralist.append(tmplist)
        return self.paralist

    def writeExcel(self, excelFile,theresult, resulttext, theRow, theCol=7):
        writeE = openpyxl.load_workbook(excelFile)
        try:
            wsheet = writeE.get_sheet_by_name('case')
            wsheet.cell(row=theRow+1, column=theCol + 1).value = resulttext
            wsheet.cell(row=theRow+1, column=theCol + 2).value = theresult

            writeE.save(excelFile)
            print('write result success!')
        except Exception as e :
            raise
        finally:
            pass

#
# if __name__ == '__main__':
#     filename = 'yk_test.xlsx'
#     oe = OptionExcel(filename)
#     caselist = oe.getExcelData(oe.excelFile)
#     # print(caselist)
#     text11 = '{"code":"A00000","data":{"cutlimitEnable":true}}'
#     for case in caselist :
#         caseid = case['id']
#         print(caseid)
#         oe.writeExcel(oe.excelFile,text11,caseid)

