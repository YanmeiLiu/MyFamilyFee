# coding = utf-8
import datetime

import openpyxl
import time

import os

import readConfig


def CreateExcel():
    xw = openpyxl.Workbook()
    # 创建文件
    sheet = xw.active
    sheet.title = 'case'
    # 创建文件表头
    sheet['A1'] = 'case_num'
    sheet['B1'] = 'business'
    sheet['C1'] = 'is_iqiyi'
    sheet['D1'] = 'is_video_page'
    sheet['E1'] = 'IP'
    sheet['F1'] = 'albumid'
    sheet['G1'] = 'tvid'
    sheet['H1'] = 'categoryid'
    sheet['I1'] = 'qypid'
    sheet['J1'] = 'expect'
    sheet['K1'] = 'result'
    sheet['L1'] = 'PASS_OR_NOT'

    # sheet中数据内容
    business = ['video_cutlimit']
    #地区，全国+韩国
    ip = ['101.83.183.223','125.39.17.91','218.70.203.113','113.126.255.163',\
          '218.12.41.179','123.188.64.65','221.208.244.214','218.62.46.191',\
          '218.28.191.23','218.95.229.8','123.81.160.205','58.192.95.214',\
          '171.82.193.63','119.39.23.134','219.137.144.45','60.162.56.231',\
          '218.64.55.198','61.138.199.193','218.66.59.145','202.100.198.24',\
          '59.48.8.183','218.88.255.123','117.34.145.123','222.86.183.149',\
          '218.22.9.4','219.159.235.101','218.202.206.45','222.75.160.43',\
          '58.18.255.255 ','220.182.50.226','61.10.255.251','122.100.128.23',\
          '114.40.105.64','218.49.74.233','10.3.14.149']
    #视频位置，必填项，为空时参数错误B00009
    is_iqiyi = ['true','false']
    #是否播放页
    is_video_page = ['true','false']
    #频道
    categoryid = ['2']
    albumid = ['']
    tvid = ['']
    qypid = ['02032001010000000000','01010011010000000000','03032001010000000000',\
             '02032001020000000000','03032001020000000000','02033001010000000000',\
             '01010011010000000000','02022001010000000000','03022001010000000000',\
             '02022001020000000000','02023001020000000000','02000021010000000000',\
             '03000021010000000000','00000021020000000000','01010011010000000000',\
             '01010011020000000000','01010011020000000000','01010010000000000000',\
             '01010010000000000000','01012000000000000000','01012001020000000000',\
             '01012001020000000000','02042001010000000000','03042001010000000000',\
             '00002001020000000000','02020001010000000000','01014111010000000000','']
    expect = ['{"code":"A00000","data":{"cutlimitEnable":false}}',\
              '{"code":"A00000","data":{"cutlimitEnable":true}}', \
              '{"code":"B00009","data":{}}']
    # print(len(expect[0]))
    # 统计有多少种组合
    blen = len(business)
    iplen = len(ip)
    ilen = len(is_iqiyi)
    vlen = len(is_video_page)
    clen = len(categoryid)
    alen = len(albumid)
    tlen = len(tvid)
    qlen = len(qypid)
    case_num = blen * ilen * vlen * clen * qlen * iplen * alen * tlen
    print('zuhecase 总共有%d条测试用例' % (case_num))
    caselist = []
    for b in business:
        # print('business=%s'%b)
        for i in is_iqiyi:
            # print('is_iqiyi=%s'%i)
            for v in is_video_page:
                # print('is_video_page=%s'%v)
                for ipn in ip :
                    for a in albumid :
                        for t in tvid :
                            for c in categoryid:
                                # print('categoryid=%s'%c)
                                for q in qypid:
                                    # print('qypid=%s'%q)
                                    # print(n)
                                    params = {
                                        'business': b, \
                                        'is_iqiyi': i, \
                                        'is_video_page': v, \
                                        'ip':ipn,\
                                        'albumid':a,\
                                        'tvid':t,\
                                        'categoryid': c, \
                                        'qypid': q}
                                    # print(params)
                                    caselist.append(params)

    for n in range(2, case_num + 2):
        # print('prepare to write')
        sheet.cell(row=n, column=1).value = n - 1
        sheet.cell(row=n, column=2).value = caselist[n - 2]['business']
        sheet.cell(row=n, column=3).value = caselist[n - 2]['is_iqiyi']
        sheet.cell(row=n, column=4).value = caselist[n - 2]['is_video_page']
        sheet.cell(row=n, column=5).value = caselist[n - 2]['ip']
        sheet.cell(row=n, column=6).value = caselist[n - 2]['albumid']
        sheet.cell(row=n, column=7).value = caselist[n - 2]['tvid']
        sheet.cell(row=n, column=8).value = caselist[n - 2]['categoryid']
        sheet.cell(row=n, column=9).value = caselist[n - 2]['qypid']
        # print('NO. %d write success' % (n - 1))

    for n in range(2,case_num+2):
        if sheet.cell(row=n,column=2).value == '' \
                or sheet.cell(row=n,column=3).value == '' \
                or sheet.cell(row=n,column=9).value == '':
            sheet.cell(row=n, column=10).value = expect[2]
        elif sheet.cell(row=n, column=9).value == '02032001010000000000' :
            sheet.cell(row=n, column=10).value = expect[1]
        else :
            sheet.cell(row=n, column=10).value = expect[0]


    dt = time.strftime('%Y%m%d-%H%M%S', time.localtime())
    result_path = readConfig.result_path
    fileresult = os.path.join(result_path , dt+'.xlsx')
    xw.save(fileresult)
    return fileresult


if  __name__ == '__main__':
    ce = CreateExcel()









