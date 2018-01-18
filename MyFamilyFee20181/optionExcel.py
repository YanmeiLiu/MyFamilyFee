# coding = utf-8

# date :2017.12.12
# user :liuyanmei

# explain :将参数组合结果保存到excel文件中


import os
import openpyxl
import time
import readconfig
import logger

class OptionExcel(object):
    def __init__(self, result):
        self.result = result

    def CreateExcel(self):
        lg = logger.config_logger('CreateExcel')
        lg.info('开始导出数据到excel文件 ')
        xw = openpyxl.Workbook()
        # 创建文件
        sheet = xw.active
        sheet.title = 'result'
        # 创建文件表头
        sheet['A1'] = '收入/支出'
        sheet['B1'] = '姓名'
        sheet['C1'] = '支付方式'
        sheet['D1'] = '金额'
        sheet['E1'] = '用途'
        sheet['F1'] = '时间'

        for n in range(3, len(self.result) + 3):
            sheet.cell(row=n - 1, column=1).value = self.result[n - 3]['shouzhi']
            sheet.cell(row=n - 1, column=2).value = self.result[n - 3]['name']
            sheet.cell(row=n - 1, column=3).value = self.result[n - 3]['payname']
            sheet.cell(row=n - 1, column=4).value = self.result[n - 3]['money']
            sheet.cell(row=n - 1, column=5).value = self.result[n - 3]['dowhat']
            sheet.cell(row=n - 1, column=6).value = self.result[n - 3]['createtime']

        dt = time.strftime('%Y%m%d-%H%M%S', time.localtime())
        result_path = readconfig.caseresult_path
        fileresult = os.path.join(result_path, dt + '.xlsx')
        xw.save(fileresult)
        lg.info('导出数据到以下文件：%s'%fileresult)
        #监测进程是否完成判断是否导出成功？

